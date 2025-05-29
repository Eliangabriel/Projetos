import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, NamedStyle
from copy import copy
from datetime import datetime
import os
from openpyxl.styles import PatternFill

# Função para abrir a janela de seleção de arquivo
def selecionar_arquivo():
    def abrir_dialogo():
        global file_path
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            label_arquivo.config(text=f"Arquivo selecionado: {file_path}")
            btn_iniciar.config(state=tk.NORMAL)  # Ativa o botão "Iniciar Processo"

    def iniciar_processo():
        if file_path:
            editar_planilha(file_path)
            messagebox.showinfo("Processo", "Processo concluído!")  # Mensagem de conclusão

    # Cria a janela principal
    root = tk.Tk()
    root.title("Selecionar Arquivo")
    root.geometry("600x400")
    
    # Aplicar o tema 'clam'
    style = ttk.Style()
    style.theme_use('clam')

    # Adiciona um frame para centralizar os widgets
    frame = ttk.Frame(root, padding="10")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    # Cria e adiciona um botão para abrir o diálogo de seleção de arquivo
    btn_selecionar = ttk.Button(frame, text="Selecionar Arquivo", command=abrir_dialogo)
    btn_selecionar.grid(row=1, column=0, pady=10)

    # Adiciona um label para mostrar o arquivo selecionado
    label_arquivo = ttk.Label(frame, text="Nenhum arquivo selecionado")
    label_arquivo.grid(row=2, column=0, pady=10)

    # Cria e adiciona um botão para iniciar o processo
    btn_iniciar = ttk.Button(frame, text="Iniciar Processo", command=iniciar_processo, state=tk.DISABLED)
    btn_iniciar.grid(row=3, column=0, pady=10)

    # Inicia o loop principal da aplicação
    root.mainloop()

# Função para aplicar filtro na planilha
def aplicar_filtro(ws):
    ws.auto_filter.ref = ws.dimensions

# Função para remover filtro de uma planilha
def remover_filtro(ws):
    if ws.auto_filter.ref:
        ws.auto_filter.ref = None

# Função principal para ler e editar a planilha
def editar_planilha(file_path):
    # Carregando a planilha original para obter o nome das sheets
    wb_origem = load_workbook(file_path)

    # Selecionando automaticamente a opção 4 ("REQ PORTAIS...") 
    escolha = "4"

    # Obtendo a planilha selecionada
    ws_base = wb_origem[wb_origem.sheetnames[int(escolha) - 1]]

    # Criando um novo arquivo Excel com duas abas
    wb_novo = Workbook()
    ws_novo1 = wb_novo.active
    ws_novo1.title = "Planilha1"
    ws_novo2 = wb_novo.create_sheet(title="Planilha2")

    # Copiando a formatação da planilha original para as novas abas
    for row in ws_base.iter_rows(min_row=1, max_row=ws_base.max_row, min_col=1, max_col=ws_base.max_column):
        for cell in row:
            new_cell1 = ws_novo1[cell.coordinate]
            new_cell2 = ws_novo2[cell.coordinate]
            
            for new_cell in [new_cell1, new_cell2]:
                new_cell.value = cell.value
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill) if cell.fill else PatternFill()  # Copiando cor de preenchimento se existir
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
                new_cell.protection = copy(cell.protection)

            # Copiando a cor de preenchimento das células das colunas B e C
            if cell.column == column_index_from_string('B') or cell.column == column_index_from_string('C'):
                if cell.fill and cell.fill.fgColor:
                    new_cell1.fill = copy(cell.fill)
                    new_cell2.fill = copy(cell.fill)

    # Excluindo a coluna 'J' da Planilha1 e alinhando as colunas subsequentes à esquerda
    ws_novo1.delete_cols(column_index_from_string('J'))
    for col in ws_novo1.iter_cols(min_col=column_index_from_string('J')):
        for cell in col:
            cell.alignment = copy(ws_base[cell.coordinate].alignment)
            cell.value = ws_base[cell.coordinate].value

    # Copiando os dados e a formatação da coluna O da planilha base para a Planilha1
    for row in ws_base.iter_rows(min_row=1, max_row=ws_base.max_row, min_col=column_index_from_string('O'), max_col=column_index_from_string('O')):
        for cell in row:
            new_cell = ws_novo1[cell.coordinate]
            new_cell.value = cell.value
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill) if cell.fill else PatternFill()  # Copiando cor de preenchimento se existir
            new_cell.number_format = copy(cell.number_format)
            new_cell.alignment = copy(cell.alignment)
            new_cell.protection = copy(cell.protection)

    # Alinhando as colunas N e O à esquerda na Planilha1
    for col in ws_novo1.iter_cols(min_col=column_index_from_string('N'), max_col=column_index_from_string('O')):
        for cell in col:
            cell.alignment = Alignment(horizontal="left")

    # Aplicando a quebra de texto em todas as células da Planilha1
    for row in ws_novo1.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Removendo a quebra de texto de todas as células, exceto da linha 1
    for row in ws_novo1.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=False)

    # Centralizando e alinhando no meio a linha 1 da Planilha1
    for cell in ws_novo1[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Centralizando a formatação da coluna A até a coluna L, a partir da linha 2
    for row in ws_novo1.iter_rows(min_row=2, min_col=column_index_from_string('A'), max_col=column_index_from_string('L')):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Aplicando a quebra de texto em todas as células da Planilha2
    for row in ws_novo2.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    for row in ws_novo2.iter_rows(min_col=column_index_from_string('A'), max_col=column_index_from_string('L')):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)       
 

    # Excluindo as colunas especificadas da Planilha1
    colunas_a_remover_planilha1 = ['J']
    for coluna in reversed(colunas_a_remover_planilha1):
        ws_novo1.delete_cols(column_index_from_string(coluna))

    # Excluindo as colunas especificadas da Planilha2
    colunas_a_remover_planilha2 = ['G', 'J', 'M', 'N', 'O']
    for coluna in reversed(colunas_a_remover_planilha2):
        ws_novo2.delete_cols(column_index_from_string(coluna))

    # Aplicando filtro nas novas abas
    aplicar_filtro(ws_novo1)
    aplicar_filtro(ws_novo2)

    # Definindo as dimensões das células da Planilha1 conforme especificado
    ws_novo1.row_dimensions[1].height = 31.5
    colunas_largura = {
        'A': 12,
        'B': 19,
        'C': 19,
        'D': 22,
        'E': 11,
        'F': 30,
        'G': 30,
        'H': 23,
        'I': 20,
        'J': 12,
        'K': 12,
        'L': 21,
        'M': 25,
        'N': 47
    }
    for col, largura in colunas_largura.items():
        ws_novo1.column_dimensions[col].width = largura

    ws_novo2.row_dimensions[1].height = 31.5
    colunas_largura = {
        'A': 12,
        'B': 19,
        'C': 19,
        'D': 16,
        'E': 11,
        'F': 37,
        'G': 21,
        'H': 16,
        'I': 12,
        'J': 12
    }
    for col, largura in colunas_largura.items():
        ws_novo2.column_dimensions[col].width = largura   

    # Definindo um estilo de data para o formato dd/mm/aaaa
    date_style = NamedStyle(name='date_style')
    date_style.number_format = 'DD/MM/YYYY'
    date_style.alignment = Alignment(horizontal='center', vertical='center')

    # Aplicando o estilo de data às células com datas na Planilha1
    for row in ws_novo1.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.style = date_style
                
    # Aplicando o estilo de data às células com datas na Planilha2
    for row in ws_novo2.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.style = date_style
    
    # Salvando o arquivo com o nome especificado
    data_atual = datetime.now().strftime("%d%m%Y")
    nome_arquivo = f"REQ_ABERTAS_STATUS_{data_atual}.xlsx"
    caminho_completo = os.path.join(os.path.dirname(file_path), nome_arquivo)
    wb_novo.save(caminho_completo)

    # Mensagem de conclusão no console
    print(f"A edição da planilha foi concluída e o arquivo foi salvo como: {nome_arquivo}")

# Chamada da função principal
if __name__ == "__main__":
    selecionar_arquivo()

